#!/usr/bin/env python3
"""
Convert Ranking.xlsx to data.csv format for the ranking system.

The Ranking.xlsx format has:
- Row 1: Header with evaluator seat numbers
- Rows 2-23: Seat numbers (25-46) with ratings from each evaluator
- Last column: Comments (銳評)
- Last row: Password (密碼)

The data.csv format needs:
- Row 1: Header with seat numbers (評價者,25號,26號,...)
- Rows 2-11: Evaluators with ratings for each seat
- Row 12: Comments for each seat (銳評)
- Row 13: Password (密碼)
"""

from openpyxl import load_workbook
import sys

def convert_xlsx_to_csv(xlsx_file='Ranking.xlsx', csv_file='data.csv'):
    # Load the Excel file
    wb = load_workbook(xlsx_file)
    ws = wb.active
    
    # Read all data
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    # Get headers from row 1
    headers = data[0]
    comment_col_idx = headers.index('銳評') if '銳評' in headers else -1
    
    # Extract seat numbers (25-46) and their data
    seats_data = {}
    password = None
    
    for i in range(1, len(data)):
        row = data[i]
        seat_num = row[0]
        
        if seat_num == '密碼':
            # Store password
            password = row[1] if len(row) > 1 else None
            continue
        
        if isinstance(seat_num, int) and 25 <= seat_num <= 46:
            # Get ratings (exclude first column and comment column)
            ratings_end = comment_col_idx if comment_col_idx > 0 else len(row)
            ratings = [row[j] if row[j] is not None else 0 for j in range(1, ratings_end)]
            
            # Get comment
            comment = row[comment_col_idx] if comment_col_idx >= 0 and comment_col_idx < len(row) else ''
            if comment is None:
                comment = ''
            
            seats_data[seat_num] = {
                'ratings': ratings,
                'comment': str(comment)
            }
    
    # Create CSV content
    csv_lines = []
    
    # Header row
    seat_numbers = sorted(seats_data.keys())
    header_line = '評價者,' + ','.join([f'{seat}號' for seat in seat_numbers])
    csv_lines.append(header_line)
    
    # Find number of evaluators
    num_evaluators = len(seats_data[seat_numbers[0]]['ratings']) if seat_numbers else 0
    
    # Rating rows (transpose the data)
    for i in range(num_evaluators):
        evaluator_name = f'評價者{i+1}'
        ratings = [str(int(seats_data[seat]['ratings'][i])) if i < len(seats_data[seat]['ratings']) else '0' 
                   for seat in seat_numbers]
        csv_lines.append(evaluator_name + ',' + ','.join(ratings))
    
    # Comments row
    comments = [seats_data[seat]['comment'] for seat in seat_numbers]
    csv_lines.append('銳評,' + ','.join(comments))
    
    # Password row
    if password:
        csv_lines.append(f'密碼,{password}')
    
    # Write to CSV file
    with open(csv_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(csv_lines))
    
    print(f'Successfully converted {xlsx_file} to {csv_file}')
    print(f'- Seat numbers: {seat_numbers[0]}-{seat_numbers[-1]}')
    print(f'- Number of evaluators: {num_evaluators}')
    print(f'- Password: {"Set" if password else "Not set"}')

if __name__ == '__main__':
    convert_xlsx_to_csv()
